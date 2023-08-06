#ATTENDANCE TRACKING SYSTEM - BY - PYTHONIC DATAMINDS
#-------------------------------------------------------------------------#
import pandas as pd                                                         #pandas library 
import openpyxl as op                                                       #open python excel library
print("______________________ATTENDANCE SYSTEM________________________")

                            #STORED NAMES OF STUDENTS#
#_________________________________________________________________________#

data_file = pd.read_excel("data.xlsx")                                       #read exist data file
wb = op.load_workbook("data.xlsx")                                           #open the workbook
sh = wb.active                                                               #load the sheet
print(data_file)                                                             #print the exist file

print("-----------------------------------------------------------")
#________________DISPLAY THE LIST AND ADD NEW STUDENT ___________#
#------------------------------------------------------------------#
#new = []
num = 1                                                                      #intialize num for y as 1 
a = (input("Any New Student, press '(y)' :  ".format(num)))                  #user input for new student
if a.lower() == 'y':                                                         #conditon for yes
    for i in range(num):                                                     #loop for add new student
        c = sh.cell(row=7, column=1)                                         #fetch roll_no col
        c.value = input("Enter roll_no : ")                                  #user input for roll_no
        c = sh.cell(row=7, column=2)                                         #fetch name col
        c.value = input("Enter Student name :" )                             #user input for name
        wb.save('data1.xlsx')                                                #save the changes in new excel file
    print("Updating ...")                                                    #printing 
else:                                                                        #else condition 
    print("No New student, Continue... ")                                    #printing

new_file = pd.read_excel("data1.xlsx")                                       #read new data file
print(new_file)                                                              #print new file
#------------------------------------------------------------------------#
                           #ATTENDANCE MARKING#
#_________________________________________________________________________#
print("---------------MARK ATTENDANCE-----------------------------")         
x = (input("Enter Student ROLL_NO : "))         
num_two = 1                                                                  #intialize num_two for name as 1 
name = (input("'(p)' for Present (n) for Absent  : ".format(num_two)))       #user input to mark attendance
if name.lower() == 'p':                                                      #if condition P for present
    print('Attendace marked')                                                #printing
else:                                                                        #else condition for absent
    print("Absent")                                                          #printing


